"""
Export service — generates PDF and Excel files for list endpoints.

Uses reportlab for PDFs and openpyxl for Excel. Returns raw bytes so endpoints
can wrap them in a StreamingResponse without touching the filesystem.
"""
from datetime import datetime
from io import BytesIO
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.models.company import Company
from app.models.deal_registration import DealRegistration
from app.models.opportunity import Opportunity

BRAND_COLOR = colors.HexColor("#1a237e")
HEADER_BG = "1A237E"  # openpyxl expects RGB hex without '#'


# ------------------------------ PDF helpers ----------------------------------

def _make_pdf_doc(buf: BytesIO, title: str) -> SimpleDocTemplate:
    return SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=title,
        author="Extravis Partner Portal",
    )


def _pdf_header_elements(title: str, subtitle: str | None = None) -> list:
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle(
        "PortalH1",
        parent=styles["Heading1"],
        textColor=BRAND_COLOR,
        fontSize=18,
        spaceAfter=4,
    )
    meta = ParagraphStyle(
        "PortalMeta",
        parent=styles["Normal"],
        textColor=colors.grey,
        fontSize=9,
    )
    generated = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    elements: list = [Paragraph(title, h1)]
    if subtitle:
        elements.append(Paragraph(subtitle, meta))
    elements.append(Paragraph(f"Generated {generated}", meta))
    elements.append(Spacer(1, 8))
    return elements


def _pdf_table(headers: Sequence[str], rows: Sequence[Sequence[str]], col_widths: Sequence[float]) -> Table:
    cell_style = ParagraphStyle(
        "Cell",
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        wordWrap="CJK",
    )
    header_style = ParagraphStyle(
        "CellHeader",
        fontName="Helvetica-Bold",
        fontSize=9,
        textColor=colors.white,
        leading=11,
    )

    data = [[Paragraph(str(h), header_style) for h in headers]]
    for row in rows:
        data.append([Paragraph(str(cell) if cell is not None else "", cell_style) for cell in row])

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), BRAND_COLOR),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f6fb")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


# ------------------------------ XLSX helpers ---------------------------------

def _xlsx_write_header(ws, headers: Sequence[str]) -> None:
    header_font = Font(color="FFFFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _xlsx_autosize(ws, headers: Sequence[str], rows: Sequence[Sequence[object]]) -> None:
    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            if idx - 1 < len(row):
                val = row[idx - 1]
                if val is not None:
                    max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 42)


def _xlsx_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ------------------------------ Opportunities --------------------------------

OPP_HEADERS = [
    "ID", "Name", "Customer", "Company", "Country", "Region",
    "Worth (USD)", "Closing Date", "Status", "Submitted By", "Created",
]


def _opportunity_row(opp: Opportunity) -> list[str]:
    return [
        opp.id,
        opp.name,
        opp.customer_name,
        opp.company.name if opp.company else "",
        opp.country,
        opp.region,
        f"{float(opp.worth):,.2f}" if opp.worth is not None else "",
        opp.closing_date.strftime("%Y-%m-%d") if opp.closing_date else "",
        opp.status.value.replace("_", " ").title() if opp.status else "",
        opp.submitted_by_user.full_name if opp.submitted_by_user else "",
        opp.created_at.strftime("%Y-%m-%d") if opp.created_at else "",
    ]


def build_opportunity_pdf(opportunities: Iterable[Opportunity], subtitle: str | None = None) -> bytes:
    buf = BytesIO()
    doc = _make_pdf_doc(buf, "Opportunities Report")
    elements = _pdf_header_elements("Opportunities Report", subtitle)

    rows = [_opportunity_row(opp) for opp in opportunities]
    if not rows:
        styles = getSampleStyleSheet()
        elements.append(Paragraph("No opportunities match the current filters.", styles["Italic"]))
    else:
        col_widths = [14 * mm, 44 * mm, 40 * mm, 40 * mm, 22 * mm, 22 * mm, 24 * mm, 22 * mm, 24 * mm, 30 * mm, 22 * mm]
        elements.append(_pdf_table(OPP_HEADERS, rows, col_widths))

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


def build_opportunity_xlsx(opportunities: Iterable[Opportunity]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunities"

    _xlsx_write_header(ws, OPP_HEADERS)
    rows = [_opportunity_row(opp) for opp in opportunities]
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    _xlsx_autosize(ws, OPP_HEADERS, rows)
    ws.freeze_panes = "A2"
    return _xlsx_to_bytes(wb)


# ------------------------------ Deals ----------------------------------------

DEAL_HEADERS = [
    "ID", "Customer", "Company", "Description", "Value (USD)", "Expected Close",
    "Status", "Exclusivity Start", "Exclusivity End", "Created",
]


def _deal_row(deal: DealRegistration) -> list[str]:
    return [
        deal.id,
        deal.customer_name,
        deal.company.name if deal.company else "",
        (deal.deal_description or "")[:200],
        f"{float(deal.estimated_value):,.2f}" if deal.estimated_value is not None else "",
        deal.expected_close_date.strftime("%Y-%m-%d") if deal.expected_close_date else "",
        deal.status.value.title() if deal.status else "",
        deal.exclusivity_start.strftime("%Y-%m-%d") if deal.exclusivity_start else "",
        deal.exclusivity_end.strftime("%Y-%m-%d") if deal.exclusivity_end else "",
        deal.created_at.strftime("%Y-%m-%d") if deal.created_at else "",
    ]


def build_deal_pdf(deals: Iterable[DealRegistration], subtitle: str | None = None) -> bytes:
    buf = BytesIO()
    doc = _make_pdf_doc(buf, "Deal Registrations Report")
    elements = _pdf_header_elements("Deal Registrations Report", subtitle)

    rows = [_deal_row(d) for d in deals]
    if not rows:
        styles = getSampleStyleSheet()
        elements.append(Paragraph("No deals match the current filters.", styles["Italic"]))
    else:
        col_widths = [14 * mm, 40 * mm, 40 * mm, 60 * mm, 24 * mm, 24 * mm, 22 * mm, 22 * mm, 22 * mm, 22 * mm]
        elements.append(_pdf_table(DEAL_HEADERS, rows, col_widths))

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


def build_deal_xlsx(deals: Iterable[DealRegistration]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Deals"

    _xlsx_write_header(ws, DEAL_HEADERS)
    rows = [_deal_row(d) for d in deals]
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    _xlsx_autosize(ws, DEAL_HEADERS, rows)
    ws.freeze_panes = "A2"
    return _xlsx_to_bytes(wb)


# ------------------------------ Companies ------------------------------------

COMPANY_HEADERS = [
    "ID", "Name", "Country", "Region", "City", "Industry",
    "Contact Email", "Tier", "Status", "Created",
]


def _company_row(c: Company) -> list[str]:
    return [
        c.id,
        c.name,
        c.country,
        c.region,
        c.city,
        c.industry,
        c.contact_email,
        c.tier.value.title() if c.tier else "",
        c.status.value.title() if c.status else "",
        c.created_at.strftime("%Y-%m-%d") if c.created_at else "",
    ]


def build_company_xlsx(companies: Iterable[Company]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"

    _xlsx_write_header(ws, COMPANY_HEADERS)
    rows = [_company_row(c) for c in companies]
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    _xlsx_autosize(ws, COMPANY_HEADERS, rows)
    ws.freeze_panes = "A2"
    return _xlsx_to_bytes(wb)


def build_company_pdf(companies: Iterable[Company], subtitle: str | None = None) -> bytes:
    buf = BytesIO()
    doc = _make_pdf_doc(buf, "Partner Companies Report")
    elements = _pdf_header_elements("Partner Companies Report", subtitle)

    rows = [_company_row(c) for c in companies]
    if not rows:
        styles = getSampleStyleSheet()
        elements.append(Paragraph("No companies match the current filters.", styles["Italic"]))
    else:
        col_widths = [14 * mm, 48 * mm, 24 * mm, 24 * mm, 24 * mm, 38 * mm, 48 * mm, 20 * mm, 22 * mm, 22 * mm]
        elements.append(_pdf_table(COMPANY_HEADERS, rows, col_widths))

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()
