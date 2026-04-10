import io
from fastapi import APIRouter, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook, load_workbook

from app.core.database import get_db
from app.core.deps import get_current_admin
from app.models.user import User, UserRole
from app.models.company import Company
from app.core.exceptions import BadRequestException
from app.utils.audit import write_audit_log

router = APIRouter(prefix="/companies", tags=["Companies"])

EXPECTED_HEADERS = [
    "Company Name",
    "Country",
    "City",
    "Industry",
    "Contact Email",
    "Channel Manager Email",
]


@router.post("/bulk-import", status_code=200)
async def bulk_import_companies(
    file: UploadFile = File(...),
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise BadRequestException(
            code="INVALID_FILE_TYPE",
            message="Only .xlsx files are accepted",
        )

    contents = await file.read()
    try:
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True)
    except Exception:
        raise BadRequestException(
            code="INVALID_FILE",
            message="Could not parse the uploaded file as a valid Excel workbook",
        )

    ws = wb.active
    if ws is None:
        raise BadRequestException(code="EMPTY_FILE", message="The workbook has no active sheet")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise BadRequestException(
            code="EMPTY_FILE",
            message="The file must contain a header row and at least one data row",
        )

    header = [str(h).strip() if h else "" for h in rows[0]]
    for expected in EXPECTED_HEADERS:
        if expected not in header:
            raise BadRequestException(
                code="INVALID_HEADERS",
                message=f"Missing required column: {expected}",
            )

    col_idx = {name: header.index(name) for name in EXPECTED_HEADERS}

    succeeded = 0
    failed = []

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            def cell(name: str) -> str:
                val = row[col_idx[name]]
                if val is None:
                    raise ValueError(f"{name} is required")
                return str(val).strip()

            company_name = cell("Company Name")
            country = cell("Country")
            city = cell("City")
            industry = cell("Industry")
            contact_email = cell("Contact Email")
            cm_email = cell("Channel Manager Email")

            # Look up channel manager
            cm_result = await db.execute(
                select(User).where(
                    User.email == cm_email,
                    User.role == UserRole.ADMIN,
                    User.deleted_at.is_(None),
                )
            )
            channel_manager = cm_result.scalar_one_or_none()
            if not channel_manager:
                raise ValueError(f"Channel manager not found or not an admin: {cm_email}")

            company = Company(
                name=company_name,
                country=country,
                region=country,  # default region to country
                city=city,
                industry=industry,
                contact_email=contact_email,
                channel_manager_id=channel_manager.id,
            )
            db.add(company)
            await db.flush()

            await write_audit_log(
                db, admin.id, "CREATE", "company", company.id,
                {"name": company_name, "source": "bulk_import"},
            )

            succeeded += 1
        except Exception as e:
            failed.append({"row": row_num, "error": str(e)})

    await db.commit()

    return {
        "processed": len(rows) - 1,
        "succeeded": succeeded,
        "failed": failed,
    }


@router.get("/bulk-import-template", status_code=200)
async def download_bulk_import_template(
    admin: User = Depends(get_current_admin),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"

    ws.append(EXPECTED_HEADERS)
    ws.append([
        "Acme Corp",
        "United States",
        "New York",
        "Technology",
        "contact@acme.com",
        "admin@extravis.com",
    ])

    # Auto-size columns for readability
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=company_import_template.xlsx"},
    )
